#  ===========================
# -*- coding:utf-8 -*-
# Time :2021/9/11 15:54
# Author :A0025-江苏-小丹
# QQ:915155536
# File :practice_xlsx.py
#  ===========================

import openpyxl

# 创建工作簿
# my_workbook = openpyxl.Workbook()
# my_workbook.save('1.xlsx')

# 打开已有工作簿
my_workbook = openpyxl.load_workbook('1.xlsx')

# 创建sheet页
# my_sheet1=my_workbook.create_sheet('my_sheet1',0)
# my_sheet2=my_workbook.create_sheet('my_sheet2')
# sheet页改名
# my_workbook.worksheets[0].title='test'
# 删除sheet页,remove()
# my_workbook.remove(my_workbook.worksheets[0])

# 根据名字获取sheet表
my_sheet = my_workbook['my_sheet1']
# print(my_sheet.title)

# 获取所有sheet表名
# print(my_workbook.sheetnames)

# 获取所有sheet表对象
# for sheet in :
#     prmy_workbookint(sheet)
# print(my_workbook.worksheets)

# #获取最大行、最大列
# print(my_sheet.max_row)
# print(my_sheet.max_column)
# 获取单元格
# print(my_sheet.cell(row=2,column=1))
# print(my_sheet['A5'])
# print(my_sheet['A1':'B5'])
# print(my_sheet[2]) #获取第二行的单元格
# 按行获取单元格
# for i in my_sheet.rows:
#     print(i)
# 按列获取单元格
# for i in my_sheet.columns:
#     print(i)

# 按行获取单元格的值,一行一个元组
# for i in my_sheet.values:
#     print(i)
# 获取单元格的值
# print(my_sheet.cell(row=2,column=1).value)

# 单元格赋值
# my_sheet['A6']='测试111'
# my_sheet.cell(row=6,column=2,value='测试222')
# my_sheet.cell(row=6,column=3).value='测试333'
# print(my_sheet['A6'].value)
# print(my_sheet['B6'].value)
# print(my_sheet['C6'].value)

# my_sheet.append([1, 2, 3])

my_sheet.append({'a':'zhang1','b':18})

my_workbook.save('1.xlsx')
