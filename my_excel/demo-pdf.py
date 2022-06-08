#  ===========================
# -*- coding:utf-8 -*-
# Time :2022/3/14 14:01
# Author :黄丹丹
# QQ:915155536
# File :demo1.py
#  ===========================
import openpyxl

wk = openpyxl.load_workbook('1.xlsx')
# print(wk.sheetnames)  # ['my_sheet1', 'Sheet', 'my_sheet2']

sheet1=wk.active
# print(sheet1)
# print(wk.sheetnames)
# 获取sheet页
my_sheet = wk['my_sheet1']
# print(my_sheet)  # <Worksheet "my_sheet1">

# 查询表中有数据的有几行几列
# print(my_sheet.dimensions)

# 获取表格中的值
# 第一种 sheet["A1"]方式
cell1 = my_sheet['A2']
# print(cell1.value)

# 第二种 sheet.cell(row=, column=)方式
cell2 = my_sheet.cell(row=1, column=2)
# print(cell2.value)

#获取某个单元格的行数、列数、坐标
# print(cell1.row)
# print(cell1.column)
# print(cell1.coordinate)

# 获取一系列格子
# 1) sheet[]方式
cell3=my_sheet['A1:B4']
# print(cell3)
# for i in cell3:
#     for j in i:
#         print(j.value)

# sheet["A"] --- 获取 A 列的数据
# sheet["A:C"] --- 获取 A,B,C 三列的数据
# sheet[5] --- 只获取第 5 行的数据

# cell4=my_sheet['A:C']
# for i in cell4:
#     print(i)
# print(cell4)

# cell5=my_sheet[2]
# for i in cell5:
#     print(i.value,type(i.value))
# print(cell5)

#② .iter_rows()方式

# for i in my_sheet.iter_rows(min_row=1,max_row=2,min_col=2,max_col=3):
#     for j in i:
#         print(j.value)

# ③获取所有行： sheet.rows
#
# for i in my_sheet.rows:
#     for j in i :
#         print(j.value)

# for i in my_sheet.columns:
#     for j in i :
#         print(j.value)

#① 向某个格子中写入内容并保存
# my_sheet['C4']='1133@qq.com'
# wk.save('1.xlsx')
# print(my_sheet['C4'].value)

# ② .append()：向表格中插入行数据(很有用)
# l1=[1,2,3,4,5,6,7]
# my_sheet.append(l1)
# wk.save('1.xlsx')

# ⑤ .delete_rows()和.delete_cols()：删除行和列
# * .delete_rows(idx=数字编号, amount=要删除的行数)
# my_sheet.delete_rows(idx=6,amount=2)
# wk.save('2.xlsx')

# #创建新sheet
# wk.create_sheet('new_sheet1')
# wk.save('1.xlsx')

#remove()：删除某个 sheet 表
# print(wk.sheetnames)
# wk.remove(wk['new_sheet1'])
# wk.save('1.xlsx')

# sheet.title：修改 sheet 表的名称
# * .title = "新的 sheet 表名"
# my_sheet.title='my_sheet1'
# wk.save('1.xlsx')

#创建新表
wk_new=openpyxl.Workbook()
wk_new['Sheet'].title='test'
wk_new.save('3.xlsx')

